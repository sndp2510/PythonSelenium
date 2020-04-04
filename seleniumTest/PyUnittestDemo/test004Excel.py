from selenium import webdriver
from selenium .webdriver.common.by import By
from selenium.webdriver.chrome.options import Options

import unittest
import os as system, time

from openpyxl import Workbook
from openpyxl import load_workbook


class Login(unittest.TestCase):
    
    def setUp(self):
        #get the current directory absolute path
        currentWorkingDirectory = system.getcwd()
        currentWorkingDirectory = currentWorkingDirectory.replace("PyUnittestDemo","")
        #path to the chromeDriver
        chromeDriverAbsolutePath = currentWorkingDirectory + "drivers\chromedriver.exe"
        print(chromeDriverAbsolutePath)
                
        options =  Options()
        options.add_argument("--start-maximized")
        self.browser = webdriver.Chrome(executable_path=chromeDriverAbsolutePath,chrome_options=options)
        self.baseUrl = "http://newtours.demoaut.com/"
           
        self.userDetailsExcelAbsolutePath = currentWorkingDirectory + "userdetail.xlsx"     
   
        
    '''def test_001(self):
        
        datasheet = self.getsheetAsListOfDictionary(self.userDetailsExcelAbsolutePath, "usersRecord")
        print(datasheet) 
        '''
        
        
    def test_login(self):
        driver=self.browser
        driver.implicitly_wait(30)
        
        
        datasheet = self.getsheetAsListOfDictionary(self.userDetailsExcelAbsolutePath, "usersRecord")
        print(datasheet)
        
        for row in datasheet:
            
            driver.get(self.baseUrl)
            print(">>", row)
            
            driver.find_element(By.NAME, "userName").send_keys(row['username'])
            driver.find_element(By.XPATH, "//input[@name='password']").send_keys(row['password'])
            driver.find_element(By.XPATH,"//input[@name='login' and @value='Login']").click()        
            
            # hard wait 
            time.sleep(3)
            headerElement = driver.find_element(By.XPATH,"//b[contains(text(),'Welcome back to') and contains(text(),'Mercury Tours!')]")
            print(headerElement.is_displayed())
            self.assertTrue(headerElement.is_displayed(), "Login Success")
            
        
                       
    def tearDown(self):
        self.browser.quit()
        
    
    #Accept the connection object and execute query and close connection
    def getsheetAsListOfDictionary (self, absExcelPath, datasheet):
        myworkbook = load_workbook(absExcelPath)
        myworksheet = myworkbook[datasheet]
        
        print(myworksheet.max_row)
        print(myworksheet.max_column)
        
        hRow = myworksheet[1]
        print("hRow>> ", hRow, len(hRow))
        
        header = []
        for cell in hRow:
            header.append(cell.value)            
                
        print("header>> ", header)
        
        list = []        
        for row in myworksheet.rows:
           
            dict = {}
            index =0            
            for cell in row:
                
                print(cell.value)
                dict[header[index]]=cell.value                
                index = index + 1
        
            list.append(dict)
            
        #print(list)
        return list
            
        
        
if __name__ == "__main__":
    unittest.main()